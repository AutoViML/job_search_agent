#!/usr/bin/env python3
"""
auto_review_jobs.py
────────────────────────────────────────────────────────────────────
Reads the latest curated_matches_*.xlsx (or .csv) file from the
outputs/ directory tree, then uses Gemini to fill in two columns:

  • "Fit for you"              — short verdict on whether the job
                                 matches the candidate's objectives
  • "Job Requirements Comments" — key requirements extracted from
                                   the live job page

Rows that already have content in either column are skipped so that
the user's manual reviews are never overwritten.

Usage
──────
  # Test on the first 3 rows only (saves to *_test_output.xlsx)
  # Always re-runs those rows fresh, regardless of prior results.
  python auto_review_jobs.py --test

  # Test on a specific number of rows
  python auto_review_jobs.py --test 5

  # Full run — processes ALL blank rows and saves to the original file
  python auto_review_jobs.py

  # Target a specific file explicitly
  python auto_review_jobs.py --file outputs/test_results/curated_matches_20260427_095915.xlsx
"""

import argparse
import glob
import json
import os
import re
import shutil
import sys
import time
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from google import genai

# ─────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────
load_dotenv()

GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
GEMINI_MODEL   = os.getenv("GEMINI_MODEL", "gemini-2.5-flash-lite")

ROOT_DIR       = Path(__file__).parent
OUTPUTS_DIR    = ROOT_DIR / "outputs"
OBJECTIVES_PATH = ROOT_DIR / "search_objectives.txt"

FIT_COL        = "Fit for you"
COMMENTS_COL   = "Job Requirements Comments"

COL_ORDER = ["Company", FIT_COL, "Role", "Location", "Salary", "Link", COMMENTS_COL]

SCRAPE_TIMEOUT = 15      # seconds for HTTP request
SLEEP_BETWEEN  = 1.5     # seconds between Gemini calls
MAX_PAGE_CHARS = 3500    # truncate scraped text to keep tokens manageable

# ─────────────────────────────────────────────
# Utility helpers
# ─────────────────────────────────────────────

# Suffixes that mark derived/backup files — excluded from auto-detection
_EXCLUDE_SUFFIXES = ("_backup", "_test_output", "_test_output_backup")


def find_latest_curated_file() -> Path:
    """Return the most-recently-modified *canonical* curated_matches file.

    Canonical means the filename stem does NOT end with a known derived
    suffix such as _backup or _test_output.
    """
    patterns = [
        str(OUTPUTS_DIR / "**" / "curated_matches_*.xlsx"),
        str(OUTPUTS_DIR / "**" / "curated_matches_*.csv"),
        str(OUTPUTS_DIR / "**" / "curated_matches_*.tsv"),
    ]
    candidates = []
    for pat in patterns:
        for p in glob.glob(pat, recursive=True):
            stem = Path(p).stem
            if not any(stem.endswith(s) for s in _EXCLUDE_SUFFIXES):
                candidates.append(p)

    if not candidates:
        sys.exit("❌  No curated_matches_* file found under outputs/. Exiting.")

    latest = max(candidates, key=os.path.getmtime)
    return Path(latest)


def scrape_job_page(url: str) -> str:
    """Fetch a job page and return its visible text (truncated)."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0 Safari/537.36"
        )
    }
    try:
        resp = requests.get(url, headers=headers, timeout=SCRAPE_TIMEOUT, allow_redirects=True)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Remove navigation / cookie banners / script text
        for tag in soup(["script", "style", "nav", "footer", "header", "noscript"]):
            tag.decompose()

        text = soup.get_text(separator=" ", strip=True)
        # Collapse whitespace
        text = re.sub(r"\s{2,}", " ", text)
        return text[:MAX_PAGE_CHARS]
    except Exception as exc:
        return f"[Could not fetch page: {exc}]"


def clean_json(text: str) -> str:
    """Strip markdown code fences from a Gemini JSON response."""
    text = text.strip()
    if text.startswith("```"):
        lines = text.splitlines()
        # drop first line (```json) and last line (```)
        lines = lines[1:] if lines[0].startswith("```") else lines
        lines = lines[:-1] if lines and lines[-1].strip() == "```" else lines
        text = "\n".join(lines).strip()
    return text


def build_prompt(objectives: str, title: str, company: str,
                 location: str, salary: str, page_text: str) -> str:
    return f"""You are a career advisor helping a new college graduate evaluate job listings.

CANDIDATE OBJECTIVES:
{objectives}

JOB LISTING:
  Title    : {title}
  Company  : {company}
  Location : {location}
  Salary   : {salary}

JOB PAGE TEXT (may be truncated):
\"\"\"{page_text}\"\"\"

TASK:
1. Decide whether this job is a good fit for the candidate given their objectives.
2. Extract the key requirements/qualifications from the job page text.

Return ONLY a JSON object with exactly these two keys:
{{
  "fit_for_you": "One or two sentences: Yes/No + reason why this job is or isn't suitable.",
  "job_requirements": "Bullet-point list of the main job requirements, each on its own line starting with '+ '."
}}

No markdown fences. No extra keys. Pure JSON only.
"""


def evaluate_with_gemini(client, objectives: str, row_data: dict) -> dict:
    """Call Gemini to evaluate one job row. Returns dict with fit and requirements."""
    page_text = scrape_job_page(row_data["link"])

    prompt = build_prompt(
        objectives=objectives,
        title=row_data.get("role", "Unknown"),
        company=row_data.get("company", "Unknown"),
        location=row_data.get("location", "Unknown"),
        salary=row_data.get("salary", "N/A"),
        page_text=page_text,
    )

    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            raw = clean_json(response.text)
            result = json.loads(raw)
            return {
                "fit_for_you": str(result.get("fit_for_you", "")).strip(),
                "job_requirements": str(result.get("job_requirements", "")).strip(),
            }
        except Exception as exc:
            if attempt < max_retries - 1:
                wait = (attempt + 1) * 3
                print(f"    ⚠️  Retry {attempt+1} after {wait}s ({exc})")
                time.sleep(wait)
            else:
                return {
                    "fit_for_you": f"[Error: {exc}]",
                    "job_requirements": "[Could not evaluate]",
                }


# ─────────────────────────────────────────────
# XLSX helpers
# ─────────────────────────────────────────────

def get_or_create_col(ws, col_name: str) -> int:
    """Return 1-based column index for col_name, creating it if absent."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if col_name in headers:
        return headers.index(col_name) + 1
    # Append at the end
    new_idx = ws.max_column + 1
    ws.cell(1, new_idx).value = col_name
    return new_idx


def clear_ai_columns_xlsx(file_path: Path) -> None:
    """Wipe the two AI-generated columns in an XLSX file so test mode
    always has rows to process, regardless of prior runs."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    fit_idx      = get_or_create_col(ws, FIT_COL)
    comments_idx = get_or_create_col(ws, COMMENTS_COL)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, fit_idx).value      = None
        ws.cell(row_idx, comments_idx).value = None
    wb.save(file_path)


def process_xlsx(file_path: Path, objectives: str, client, test_limit: int | None,
                 output_path: Path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Ensure target columns exist
    fit_col_idx      = get_or_create_col(ws, FIT_COL)
    comments_col_idx = get_or_create_col(ws, COMMENTS_COL)

    # Build header → column index map
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    def col(name):
        return headers.get(name)

    processed = 0
    skipped   = 0
    total_blank = 0

    for row_idx in range(2, ws.max_row + 1):
        fit_val      = ws.cell(row_idx, fit_col_idx).value
        comments_val = ws.cell(row_idx, comments_col_idx).value

        # Skip rows already filled in
        if fit_val or comments_val:
            skipped += 1
            continue

        total_blank += 1
        if test_limit is not None and processed >= test_limit:
            continue  # count blanks but don't process beyond limit

        link = ws.cell(row_idx, col("Link")).value if col("Link") else None
        if not link:
            print(f"  Row {row_idx}: no link, skipping.")
            continue

        row_data = {
            "role":     ws.cell(row_idx, col("Role")).value     if col("Role")     else "",
            "company":  ws.cell(row_idx, col("Company")).value  if col("Company")  else "",
            "location": ws.cell(row_idx, col("Location")).value if col("Location") else "",
            "salary":   ws.cell(row_idx, col("Salary")).value   if col("Salary")   else "N/A",
            "link":     link,
        }

        print(f"  Row {row_idx}: [{row_data['company']}] {row_data['role'][:60]}")
        result = evaluate_with_gemini(client, objectives, row_data)

        ws.cell(row_idx, fit_col_idx).value      = result["fit_for_you"]
        ws.cell(row_idx, comments_col_idx).value = result["job_requirements"]

        processed += 1
        print(f"    ✅ Fit: {result['fit_for_you'][:80]}")
        time.sleep(SLEEP_BETWEEN)

    wb.save(output_path)
    return processed, skipped, total_blank


# ─────────────────────────────────────────────
# CSV / TSV helpers
# ─────────────────────────────────────────────

def process_delimited(file_path: Path, objectives: str, client, test_limit: int | None,
                      output_path: Path, delimiter: str = ","):
    """Process a CSV or TSV file (delimiter controls which)."""
    import csv

    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        rows   = list(reader)
        fieldnames = list(reader.fieldnames or [])

    if FIT_COL not in fieldnames:
        fieldnames = [fieldnames[0], FIT_COL] + fieldnames[1:]
    if COMMENTS_COL not in fieldnames:
        fieldnames.append(COMMENTS_COL)

    processed = skipped = total_blank = 0

    for row in rows:
        fit_val      = row.get(FIT_COL,      "").strip()
        comments_val = row.get(COMMENTS_COL, "").strip()

        if fit_val or comments_val:
            skipped += 1
            continue

        total_blank += 1
        if test_limit is not None and processed >= test_limit:
            continue

        link = row.get("Link", "").strip()
        if not link:
            continue

        row_data = {
            "role":     row.get("Role", ""),
            "company":  row.get("Company", ""),
            "location": row.get("Location", ""),
            "salary":   row.get("Salary", "N/A"),
            "link":     link,
        }

        print(f"  [{row_data['company']}] {row_data['role'][:60]}")
        result = evaluate_with_gemini(client, objectives, row_data)

        row[FIT_COL]      = result["fit_for_you"]
        row[COMMENTS_COL] = result["job_requirements"]

        processed += 1
        print(f"    ✅ Fit: {result['fit_for_you'][:80]}")
        time.sleep(SLEEP_BETWEEN)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore",
                                delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)

    return processed, skipped, total_blank


def process_csv(file_path: Path, objectives: str, client, test_limit: int | None,
                output_path: Path):
    return process_delimited(file_path, objectives, client, test_limit, output_path,
                             delimiter=",")


def process_tsv(file_path: Path, objectives: str, client, test_limit: int | None,
                output_path: Path):
    return process_delimited(file_path, objectives, client, test_limit, output_path,
                             delimiter="\t")


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Auto-fill 'Fit for you' and 'Job Requirements Comments' using Gemini."
    )
    parser.add_argument(
        "--test",
        nargs="?",
        const=3,
        type=int,
        metavar="N",
        help="Test mode: process only N blank rows (default 3). "
             "Saves to *_test_output.xlsx instead of the original.",
    )
    parser.add_argument(
        "--file",
        type=str,
        default=None,
        metavar="PATH",
        help="Path to a specific curated_matches file (overrides auto-detection).",
    )
    args = parser.parse_args()

    test_mode  : bool      = args.test is not None
    test_limit : int | None = args.test if test_mode else None

    # ── Validate env ──────────────────────────────────────────────
    if not GOOGLE_API_KEY:
        sys.exit("❌  GOOGLE_API_KEY (or GEMINI_API_KEY) not found in .env")

    # ── Load objectives ───────────────────────────────────────────
    if not OBJECTIVES_PATH.exists():
        sys.exit(f"❌  {OBJECTIVES_PATH} not found.")
    objectives = OBJECTIVES_PATH.read_text(encoding="utf-8").strip()

    # ── Find input file ───────────────────────────────────────────
    if args.file:
        src = Path(args.file)
        if not src.exists():
            sys.exit(f"❌  File not found: {args.file}")
    else:
        src = find_latest_curated_file()
    suffix = src.suffix.lower()

    if test_mode:
        output_path = src.with_name(src.stem + "_test_output" + suffix)
        # Copy the original so we never touch it
        shutil.copy2(src, output_path)
        # Clear AI columns in the copy so test mode always has rows to process
        if suffix == ".xlsx":
            clear_ai_columns_xlsx(output_path)
        working_path = output_path
        print(f"🧪  TEST MODE ({test_limit} rows) | Input : {src.name}")
        print(f"                             | Output: {output_path.name}")
    else:
        # Make a backup of the original before modifying
        backup_path = src.with_name(src.stem + "_backup" + suffix)
        if not backup_path.exists():
            shutil.copy2(src, backup_path)
            print(f"💾  Backup saved → {backup_path.name}")
        working_path = src
        output_path  = src
        print(f"🚀  FULL MODE | File: {src.name}")

    print(f"🤖  Model: {GEMINI_MODEL}\n")

    client = genai.Client(api_key=GOOGLE_API_KEY)

    # ── Process ───────────────────────────────────────────────────
    if suffix == ".xlsx":
        processed, skipped, total_blank = process_xlsx(
            working_path, objectives, client, test_limit, output_path
        )
    elif suffix == ".csv":
        processed, skipped, total_blank = process_csv(
            working_path, objectives, client, test_limit, output_path
        )
    elif suffix == ".tsv":
        processed, skipped, total_blank = process_tsv(
            working_path, objectives, client, test_limit, output_path
        )
    else:
        sys.exit(f"❌  Unsupported file type: {suffix} (expected .xlsx, .csv, or .tsv)")

    # ── Summary ───────────────────────────────────────────────────
    remaining = total_blank - processed
    print()
    print("=" * 55)
    print(f"  ✅  Done!  Processed : {processed} rows")
    print(f"             Skipped   : {skipped} rows (already filled)")
    print(f"             Remaining : {remaining} blank rows (not processed)")
    print(f"  📄  Saved  → {output_path}")
    print("=" * 55)


if __name__ == "__main__":
    main()
